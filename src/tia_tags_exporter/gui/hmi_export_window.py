from __future__ import annotations

import json
import os
import tempfile
import zipfile
from dataclasses import asdict, fields, is_dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook
from PySide6 import QtCore, QtWidgets

from tia_tags_exporter.config_store import ProfileStore
from tia_tags_exporter.hmi_exporter import HmiExporter, HmiExportResult, HmiTargetInfo
from tia_tags_exporter.hmi_flatteners import flatten_export_result
from tia_tags_exporter.session import TiaSession


class HmiExportWindow(QtWidgets.QMainWindow):
    """GUI front-end that launches the Openness HMI exports."""

    def __init__(
        self, store: ProfileStore, parent: Optional[QtWidgets.QWidget] = None
    ) -> None:
        super().__init__(parent)
        self.store = store
        self._session: Optional[TiaSession] = None
        self._exporter: Optional[HmiExporter] = None
        self._targets: List[HmiTargetInfo] = []

        self.setWindowTitle("TIA HMI Export")
        self.resize(900, 600)

        central = QtWidgets.QWidget(self)
        self.setCentralWidget(central)

        layout = QtWidgets.QVBoxLayout(central)

        self.lblProject = QtWidgets.QLabel("Project: <not attached>")
        layout.addWidget(self.lblProject)

        info = QtWidgets.QLabel(
            "Use this view to export native WinCC (HMI) assets via TIA Openness.\n"
            "Each run generates Siemens AML payloads for text lists, screens, alarms, and recipes."
        )
        info.setWordWrap(True)
        layout.addWidget(info)

        format_row = QtWidgets.QHBoxLayout()
        format_row.addWidget(QtWidgets.QLabel("Output:"))
        self.cmbFormat = QtWidgets.QComboBox()
        self.cmbFormat.addItem(
            "Siemens AML folder (one folder per target)", "aml-folder"
        )
        self.cmbFormat.addItem("Siemens AML zip archive (.zip)", "zip")
        format_row.addWidget(self.cmbFormat)
        format_row.addStretch(1)
        layout.addLayout(format_row)

        self.targetList = QtWidgets.QListWidget()
        self.targetList.setSelectionMode(
            QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection
        )
        self.targetList.setMinimumHeight(180)
        layout.addWidget(self.targetList)

        self.btnRun = QtWidgets.QPushButton("Export Selected HMI Targets")
        self.btnRun.setEnabled(False)
        self.btnRun.clicked.connect(self._on_run_clicked)
        layout.addWidget(self.btnRun)

        self.btnRunFlattened = QtWidgets.QPushButton("Export and Flatten")
        self.btnRunFlattened.setEnabled(False)
        self.btnRunFlattened.clicked.connect(self._on_run_flattened_export_clicked)

        self.cmbFlattenedFormat = QtWidgets.QComboBox()
        self.cmbFlattenedFormat.addItem("Excel (.xlsx)", "xlsx")

        flattened_export_row = QtWidgets.QHBoxLayout()
        flattened_export_row.addWidget(self.btnRunFlattened)
        flattened_export_row.addWidget(QtWidgets.QLabel("Format:"))
        flattened_export_row.addWidget(self.cmbFlattenedFormat)
        flattened_export_row.addStretch(1)
        layout.addLayout(flattened_export_row)

        self.log = QtWidgets.QTextEdit(readOnly=True)
        self.log.setPlaceholderText("HMI export status and logs will appear here.")
        layout.addWidget(self.log, stretch=1)

    # ------------------------------------------------------------------
    # Public API used by MainWindow
    # ------------------------------------------------------------------
    def set_session(self, session: Optional[TiaSession]) -> None:
        self._session = session
        self._exporter = None
        self._targets = []
        self.targetList.clear()
        if not session:
            self.lblProject.setText("Project: <not attached>")
            self.btnRun.setEnabled(False)
            self.btnRunFlattened.setEnabled(False)
            return

        project_name = getattr(session, "project_name", None) or getattr(
            session, "project", None
        )
        if project_name:
            self.lblProject.setText(f"Project: {project_name}")
        else:
            self.lblProject.setText("Project: <unknown>")

        try:
            self._exporter = HmiExporter(session)
            self._targets = self._exporter.list_targets()
        except Exception as exc:  # pragma: no cover - requires Openness runtime
            self._append_log(f"Failed to enumerate HMI targets: {exc}")
            self.btnRun.setEnabled(False)
            self.btnRunFlattened.setEnabled(False)
            return

        if not self._targets:
            self._append_log("No HMI targets detected in the attached project.")
            self.btnRun.setEnabled(False)
            self.btnRunFlattened.setEnabled(False)
            return

        for target in self._targets:
            label = f"{target.software_name} ({target.device_name})"
            item = QtWidgets.QListWidgetItem(label)
            item.setData(QtCore.Qt.ItemDataRole.UserRole, target.identifier)
            self.targetList.addItem(item)
        self.btnRun.setEnabled(True)
        self.btnRunFlattened.setEnabled(True)

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------
    def _on_run_flattened_export_clicked(self) -> None:
        exporter = self._exporter
        if exporter is None:
            QtWidgets.QMessageBox.warning(
                self, "HMI Export", "Attach to a project with HMI targets first."
            )
            return

        identifiers = self._selected_identifiers()
        fmt = self.cmbFlattenedFormat.currentData()

        if fmt == "xlsx":
            out, _ = QtWidgets.QFileDialog.getSaveFileName(
                self, "Save Excel", "HMI_Flattened.xlsx", "Excel (*.xlsx)"
            )
            if not out:
                return
            out_path = Path(out)
            if out_path.suffix.lower() != ".xlsx":
                out_path = out_path.with_suffix(".xlsx")

            self._run_flattened_export(identifiers, exporter, out_path)

    def _run_flattened_export(
        self,
        identifiers: Optional[Iterable[str]],
        exporter: HmiExporter,
        out_path: Path,
    ) -> None:
        self._append_log(f"Exporting and flattening to {out_path} ...")
        QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.CursorShape.WaitCursor)
        try:
            with tempfile.TemporaryDirectory(prefix="tia-hmi-export-") as tmp_dir:
                tmp_root = Path(tmp_dir)
                results = exporter.export_targets(
                    tmp_root, target_filter=identifiers, progress=self._append_log
                )
                if not results:
                    QtWidgets.QApplication.restoreOverrideCursor()
                    QtWidgets.QMessageBox.information(
                        self, "HMI Export", "No HMI targets were exported."
                    )
                    return

                all_sheets = {}
                for result in results:
                    project_name = getattr(self._session, "project_name", None)
                    flattened = flatten_export_result(result, project_name=project_name)
                    for key, data in flattened.to_mapping().items():
                        if key not in all_sheets:
                            all_sheets[key] = []
                        all_sheets[key].extend(data)

                _write_hmi_sheets(all_sheets, out_path)

        except Exception as exc:  # pragma: no cover - requires Openness runtime
            QtWidgets.QApplication.restoreOverrideCursor()
            QtWidgets.QMessageBox.critical(self, "HMI Export", f"Export failed:\n{exc}")
            self._append_log(f"Export failed: {exc}")
            return
        QtWidgets.QApplication.restoreOverrideCursor()
        message = f"Exported and flattened HMI data to:\n{out_path}"
        self.log.append(message)
        QtWidgets.QMessageBox.information(self, "Done", message)

    def _selected_identifiers(self) -> Optional[List[str]]:
        items = self.targetList.selectedItems()
        if not items:
            return None
        return [item.data(QtCore.Qt.ItemDataRole.UserRole) for item in items]

    def _prompt_output_folder(self) -> Optional[Path]:
        directory = QtWidgets.QFileDialog.getExistingDirectory(
            self, "Choose export folder"
        )
        return Path(directory) if directory else None

    def _prompt_zip_path(self) -> Optional[Path]:
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Save AML Zip", "HMI_Exports.zip", "Zip (*.zip)"
        )
        if not file_path:
            return None
        return Path(file_path)

    def _append_log(self, message: str) -> None:
        self.log.append(message)

    def _on_run_clicked(self) -> None:
        exporter = self._exporter
        if exporter is None:
            QtWidgets.QMessageBox.warning(
                self, "HMI Export", "Attach to a project with HMI targets first."
            )
            return

        identifiers = self._selected_identifiers()
        fmt = self.cmbFormat.currentData()
        if fmt == "zip":
            self._run_zip_export(identifiers, exporter)
        else:
            self._run_folder_export(identifiers, exporter)

    def _run_folder_export(
        self, identifiers: Optional[Iterable[str]], exporter: HmiExporter
    ) -> None:
        out_dir = self._prompt_output_folder()
        if not out_dir:
            return
        out_dir = Path(out_dir)
        self._append_log(f"Exporting to {out_dir} ...")
        QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.CursorShape.WaitCursor)
        try:
            results = exporter.export_targets(
                out_dir, target_filter=identifiers, progress=self._append_log
            )
        except Exception as exc:  # pragma: no cover - requires Openness runtime
            QtWidgets.QApplication.restoreOverrideCursor()
            QtWidgets.QMessageBox.critical(self, "HMI Export", f"Export failed:\n{exc}")
            self._append_log(f"Export failed: {exc}")
            return
        QtWidgets.QApplication.restoreOverrideCursor()
        manifest_payload = _build_manifest(
            results, destination=out_dir, relative_root=out_dir
        )
        self._handle_results(results, out_dir, manifest_payload)

    def _run_zip_export(
        self, identifiers: Optional[Iterable[str]], exporter: HmiExporter
    ) -> None:
        zip_path = self._prompt_zip_path()
        if not zip_path:
            return
        zip_path = Path(zip_path)
        self._append_log("Exporting to temporary folder before zipping ...")
        QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.CursorShape.WaitCursor)
        try:
            with tempfile.TemporaryDirectory(prefix="tia-hmi-export-") as tmp_dir:
                tmp_root = Path(tmp_dir)
                results = exporter.export_targets(
                    tmp_root, target_filter=identifiers, progress=self._append_log
                )
                if not results:
                    QtWidgets.QApplication.restoreOverrideCursor()
                    QtWidgets.QMessageBox.information(
                        self, "HMI Export", "No HMI targets were exported."
                    )
                    return
                manifest_payload = _build_manifest(
                    results, destination=zip_path, relative_root=tmp_root
                )
                _pack_zip(tmp_root, zip_path)
        except Exception as exc:  # pragma: no cover - requires Openness runtime
            QtWidgets.QApplication.restoreOverrideCursor()
            QtWidgets.QMessageBox.critical(self, "HMI Export", f"Export failed:\n{exc}")
            self._append_log(f"Export failed: {exc}")
            return
        QtWidgets.QApplication.restoreOverrideCursor()
        self._handle_results(results, zip_path, manifest_payload)

    def _handle_results(
        self,
        results: List[HmiExportResult],
        destination: Path,
        manifest_payload: Optional[dict] = None,
    ) -> None:
        if not results:
            QtWidgets.QMessageBox.information(
                self, "HMI Export", "No HMI targets were exported."
            )
            return
        manifest_payload = manifest_payload or _build_manifest(
            results,
            destination=destination,
            relative_root=destination if destination.is_dir() else None,
        )
        manifest_path = _write_manifest(destination, manifest_payload)
        self._append_log(f"Manifest written to {manifest_path}")
        message = "Exported {} HMI target(s).\nOutputs stored at:\n{}".format(
            len(results), destination
        )
        QtWidgets.QMessageBox.information(self, "HMI Export", message)


# ----------------------------------------------------------------------
# Helper functions
# ----------------------------------------------------------------------


def _write_hmi_sheets(sheets: Dict[str, List[Any]], out_path: Path) -> None:
    """Write a dictionary of data to an Excel workbook with multiple sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    for sheet_name, rows in sheets.items():
        if not rows:
            continue

        ws = wb.create_sheet(sheet_name)

        first_row = rows[0]
        headers = []
        if is_dataclass(first_row) and not isinstance(first_row, type):
            headers = [field.name for field in fields(first_row)]
        elif isinstance(first_row, dict):
            headers = list(first_row.keys())

        if not headers:
            continue

        ws.append(headers)

        for row in rows:
            row_dict = {}
            if is_dataclass(row) and not isinstance(row, type):
                row_dict = asdict(row)
            elif isinstance(row, dict):
                row_dict = row

            if not row_dict:
                continue

            values = []
            for header in headers:
                val = row_dict.get(header)
                if val is None:
                    values.append("")
                elif isinstance(val, (list, dict, set)):
                    values.append(str(val))
                else:
                    values.append(val)
            ws.append(values)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


def _build_manifest(
    results: List[HmiExportResult],
    destination: Path,
    relative_root: Optional[Path],
) -> dict:
    def format_path(path: Optional[Path]) -> Optional[str]:
        if not path:
            return None
        path = Path(path)
        if relative_root:
            try:
                return str(path.relative_to(relative_root))
            except ValueError:
                pass
        return str(path)

    manifest_targets = []
    for result in results:
        manifest_targets.append(
            {
                "identifier": result.target.identifier,
                "device": result.target.device_name,
                "software": result.target.software_name,
                "text_lists": format_path(result.text_list_path),
                "screens": format_path(result.screen_path),
                "alarms": format_path(result.alarm_path),
                "recipes": format_path(result.recipe_path),
                "extras": {
                    key: format_path(value)
                    for key, value in sorted(result.extras.items())
                },
            }
        )
    return {"destination": str(destination), "targets": manifest_targets}


def _write_manifest(destination: Path, payload: dict) -> Path:
    payload_text = json.dumps(payload, indent=2)
    if destination.suffix.lower() == ".zip":
        manifest_path = destination.with_suffix(".manifest.json")
        manifest_path.write_text(payload_text, encoding="utf-8")
        with zipfile.ZipFile(destination, "a", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("manifest.json", payload_text)
        return manifest_path
    manifest_path = destination / "manifest.json"
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(payload_text, encoding="utf-8")
    return manifest_path


def _pack_zip(source: Path, zip_path: Path) -> None:
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for root, _, files in os.walk(source):
            for file in files:
                full = Path(root) / file
                arc_name = full.relative_to(source)
                zf.write(full, arcname=str(arc_name))


__all__ = ["HmiExportWindow"]
