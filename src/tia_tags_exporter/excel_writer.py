from __future__ import annotations
import csv
from pathlib import Path
from typing import Iterable, List, Optional
from openpyxl import Workbook

_DEF_HEADERS = [
    "ProjectName",
    "PLC_Name",
    "TagTable",
    "TagName",
    "DataType",
    "Address",
    "Comment",
    "Retentive",
    "Scope",
    "TagId",
]


def _row_as_list(row) -> List:
    if hasattr(row, "__dict__"):
        return [getattr(row, header, "") for header in _DEF_HEADERS]
    if isinstance(row, dict):
        return [row.get(header, "") for header in _DEF_HEADERS]
    return list(row)


def _rows_as_lists(rows: Iterable) -> List[List]:
    return [_row_as_list(r) for r in rows]


def _rows_as_csv_lists(rows: Iterable) -> List[List[str]]:
    out: List[List[str]] = []
    for values in _rows_as_lists(rows):
        row_out: List[str] = []
        for value in values:
            if isinstance(value, bool):
                row_out.append("TRUE" if value else "FALSE")
            elif value is None:
                row_out.append("")
            else:
                row_out.append(str(value))
        out.append(row_out)
    return out


def write_tags_xlsx(rows: Iterable, out_path: Path) -> None:
    rows = list(rows)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("PLC_Tags")
    ws.append(_DEF_HEADERS)
    for values in _rows_as_lists(rows):
        ws.append(values)
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        std = wb["Sheet"]
        wb.remove(std)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def write_tags_csv(rows: Iterable, out_path: Path) -> None:
    rows = list(rows)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(_DEF_HEADERS)
        for values in _rows_as_csv_lists(rows):
            writer.writerow(values)


def write_tags_google_sheets(
    rows: Iterable,
    credentials_path: Path,
    spreadsheet_title: str,
    worksheet_title: str = "PLC_Tags",
    share_with: Optional[str] = None,
) -> dict:
    rows = list(rows)
    try:
        import gspread  # type: ignore
        from google.oauth2.service_account import Credentials  # type: ignore
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise RuntimeError(
            "Google Sheets export requires the 'gspread' and 'google-auth' packages."
        ) from exc

    scope = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive.file",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(str(credentials_path), scopes=scope)
    client = gspread.authorize(creds)

    spreadsheet = client.create(spreadsheet_title)
    worksheet = spreadsheet.sheet1
    worksheet.update_title(worksheet_title)

    values = _rows_as_lists(rows)
    payload = [_DEF_HEADERS] + values
    worksheet.update("A1", payload, value_input_option="USER_ENTERED")

    if share_with:
        try:
            spreadsheet.share(share_with, perm_type="user", role="writer")
        except Exception:
            pass

    return {
        "url": spreadsheet.url,
        "spreadsheet_id": spreadsheet.id,
        "title": spreadsheet.title,
        "worksheet_title": worksheet.title,
    }

_DN_HEADERS = [
    "ProjectName",
    "DeviceName",
    "DeviceType",
    "NetworkInterface",
    "NodeAddress",
    "SubnetName",
    "IoSystemName",
]


def _dn_row_as_list(row) -> List:
    if hasattr(row, "__dict__"):
        return [getattr(row, header, "") for header in _DN_HEADERS]
    if isinstance(row, dict):
        return [row.get(header, "") for header in _DN_HEADERS]
    return list(row)


def _dn_rows_as_lists(rows: Iterable) -> List[List]:
    return [_dn_row_as_list(r) for r in rows]


def _dn_rows_as_csv_lists(rows: Iterable) -> List[List[str]]:
    out: List[List[str]] = []
    for values in _dn_rows_as_lists(rows):
        row_out: List[str] = []
        for value in values:
            if isinstance(value, bool):
                row_out.append("TRUE" if value else "FALSE")
            elif value is None:
                row_out.append("")
            else:
                row_out.append(str(value))
        out.append(row_out)
    return out


def write_devices_networks_xlsx(rows: Iterable, out_path: Path) -> None:
    rows = list(rows)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Devices_Networks")
    ws.append(_DN_HEADERS)
    for values in _dn_rows_as_lists(rows):
        ws.append(values)
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        std = wb["Sheet"]
        wb.remove(std)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def write_devices_networks_csv(rows: Iterable, out_path: Path) -> None:
    rows = list(rows)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(_DN_HEADERS)
        for values in _dn_rows_as_csv_lists(rows):
            writer.writerow(values)

def write_devices_networks_google_sheets(
    rows: Iterable,
    credentials_path: Path,
    spreadsheet_title: str,
    worksheet_title: str = "Devices_Networks",
    share_with: Optional[str] = None,
) -> dict:
    rows = list(rows)
    try:
        import gspread  # type: ignore
        from google.oauth2.service_account import Credentials  # type: ignore
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise RuntimeError(
            "Google Sheets export requires the 'gspread' and 'google-auth' packages."
        ) from exc

    scope = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive.file",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(str(credentials_path), scopes=scope)
    client = gspread.authorize(creds)

    spreadsheet = client.create(spreadsheet_title)
    worksheet = spreadsheet.sheet1
    worksheet.update_title(worksheet_title)

    values = _dn_rows_as_lists(rows)
    payload = [_DN_HEADERS] + values
    worksheet.update("A1", payload, value_input_option="USER_ENTERED")

    if share_with:
        try:
            spreadsheet.share(share_with, perm_type="user", role="writer")
        except Exception:
            pass

    return {
        "url": spreadsheet.url,
        "spreadsheet_id": spreadsheet.id,
        "title": spreadsheet.title,
        "worksheet_title": worksheet.title,
    }
