from __future__ import annotations

import csv
from dataclasses import asdict, is_dataclass
from pathlib import Path
from typing import Iterable, List, Mapping, Sequence

from openpyxl import Workbook

_BLOCK_HEADERS = [
    "ProjectName",
    "PLC_Name",
    "BlockName",
    "BlockType",
    "Language",
    "AttributeName",
    "AttributeValue",
    "InterfaceSection",
    "InterfaceName",
    "InterfaceDataType",
    "InitialValue",
    "Comment",
    "SourceReference",
]


def _row_to_mapping(row) -> Mapping[str, object]:
    if is_dataclass(row):
        return asdict(row)
    if hasattr(row, "__dict__"):
        return row.__dict__
    if isinstance(row, Mapping):
        return row
    if isinstance(row, Sequence):
        data = {}
        for idx, header in enumerate(_BLOCK_HEADERS):
            if idx < len(row):
                data[header] = row[idx]
        return data
    return {header: "" for header in _BLOCK_HEADERS}


def _coerce_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def _rows_as_lists(rows: Iterable) -> List[List[str]]:
    out: List[List[str]] = []
    for row in rows:
        mapping = _row_to_mapping(row)
        out.append([_coerce_value(mapping.get(header, "")) for header in _BLOCK_HEADERS])
    return out


def write_blocks_xlsx(rows: Iterable, out_path: Path) -> None:
    rows = list(rows)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Program_Blocks")
    ws.append(_BLOCK_HEADERS)
    for values in _rows_as_lists(rows):
        ws.append(values)
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["Sheet"])
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def write_blocks_csv(rows: Iterable, out_path: Path) -> None:
    rows = list(rows)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(_BLOCK_HEADERS)
        for values in _rows_as_lists(rows):
            writer.writerow(values)


def write_blocks_google_sheets(
    rows: Iterable,
    credentials_path: Path,
    spreadsheet_title: str,
    share_with: str | None = None,
) -> dict:
    rows = list(rows)
    try:
        import gspread  # type: ignore
        from google.oauth2.service_account import Credentials  # type: ignore
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise RuntimeError(
            "Google Sheets export requires the 'gspread' and 'google-auth' packages."
        ) from exc

    scope = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive.file",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(str(credentials_path), scopes=scope)
    client = gspread.authorize(creds)

    spreadsheet = client.create(spreadsheet_title)
    worksheet = spreadsheet.sheet1
    worksheet.update_title("Program_Blocks")

    payload = [_BLOCK_HEADERS] + _rows_as_lists(rows)
    worksheet.update("A1", payload, value_input_option="USER_ENTERED")

    if share_with:
        try:
            spreadsheet.share(share_with, perm_type="user", role="writer")
        except Exception:
            pass

    return {
        "url": spreadsheet.url,
        "spreadsheet_id": spreadsheet.id,
        "title": spreadsheet.title,
        "worksheet_title": worksheet.title,
    }
